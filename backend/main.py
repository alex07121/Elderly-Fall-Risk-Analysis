# Web dashboard address: 127.0.0.1:8000
# Interactive API control panel: 127.0.0.1:8000/docs
# username: admin_clinician
# password: password123

# Download library
from backend.libAuto import libmain
libmain()

import os
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
import hashlib
import secrets
from fastapi import Depends, FastAPI, HTTPException, status
from fastapi.responses import Response
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
import jwt
from pydantic import BaseModel, Field
# This imports code and loads the models automatically
from ml.predict import predict_fall_risk, explain_patient
import gradio as gr
import pandas as pd
import random
import json
import asyncio
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from backend.models import PatientRecord
from backend.database import Base, engine, get_db, AsyncSessionLocal

# Project paths (backend/ sits one level below the project root)
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
from sqlalchemy.ext.asyncio import AsyncSession

from fastapi.middleware.cors import CORSMiddleware
from typing import Literal, Optional

SECRET_KEY = "SUPER_SECRET_SECURITY_KEY_CHANGE_THIS_IN_PRODUCTION"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# hashlib.sha256(b"password123").hexdigest()
ADMIN_PASSWORD_HASH = "ef92b778bafe771e89245b89ecbc08a44a4e166c06659911881f383d4473e94f"

MOCK_USERS_DB = {
    "admin_clinician": {
        "username": "admin_clinician",
        "hashed_password": ADMIN_PASSWORD_HASH,
        "disabled": False,
    }
}

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

class Token(BaseModel):
    access_token: str
    token_type: str

def verify_password(plain_password: str, correct_hash: str) -> bool:
    input_hash = hashlib.sha256(plain_password.encode('utf-8')).hexdigest()
    return secrets.compare_digest(input_hash, correct_hash)

def create_access_token(data: dict, expires_delta: timedelta | None = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# Dependency validation check ensuring route protection
async def get_current_user(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except jwt.PyJWTError:
        raise credentials_exception
        
    user = MOCK_USERS_DB.get(username)
    if user is None:
        raise credentials_exception
    return user


@asynccontextmanager
async def lifespan(_app: FastAPI):
    # Startup
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    yield
    # Shutdown
    await engine.dispose()

# Initialize the API
app = FastAPI(title="Fall Risk Assessment API", lifespan=lifespan)

# Define the 10 inputs exactly as requested in the README
class PatientData(BaseModel):
    sex: Literal['M', 'F']
    age: int = Field(..., ge=60, le=100)
    night_bed_exits: int = Field(..., ge=0, le=8)
    night_activity_duration_min: float = Field(..., ge=0, le=120)
    past_falls: int = Field(..., ge=0, le=5)
    mobility_score: int = Field(..., ge=1, le=10)
    high_risk_medication: int = Field(..., ge=0, le=1)
    cognitive_impairment: int = Field(..., ge=0, le=2)
    polypharmacy_count: int = Field(..., ge=0, le=14)
    orthostatic_hypotension: int = Field(..., ge=0, le=1)
    tug_seconds: float = Field(..., ge=8.0, le=31.9)
    # Extended fall-detail fields (NOT model inputs; kept for clinical context & dashboard display)
    days_since_last_fall: Optional[int] = Field(None, ge=0)
    syncopal_fall: int = Field(0, ge=0, le=1)
    fall_cluster_30d: int = Field(0, ge=0, le=1)
    resident_id: Optional[str] = None  # optional, links repeated assessments of the same resident


# ---------- 60-74 岁带跌倒风险建议（特征驱动，科学依据见各条注释） ----------
# 优先级: 1 = 立即處理, 2 = 本週做, 3 = 例行
# 依据: TUG (Podsiadlo & Richardson 1991) / Morse Fall Scale / Beers 2023 /
#       STOPPFall / CDC STEADI / Otago Exercise Programme / MMSE

SUGGESTION_PRIORITIES = {
    1: "Act now",
    2: "This week",
    3: "Routine",
}


def _suggest_60_74(features: dict) -> list:
    """60-74 岁带动态建议：数值嵌入 + 组合推导 + 一句话精简。

    三层：_describe 逐特征生成含数值的 label/action → _apply_combos 组合吸收 →
    _finalize 过滤排序限 6 条。对外只保留 {feature, label, action, priority}。
    """
    def num(v):
        """数值显示规范化：55.0 → 55，11.2 → 11.2。"""
        try:
            f = float(v)
            return int(f) if f == int(f) else f
        except (TypeError, ValueError):
            return v

    items = []

    def add(feature, label, action, priority, topic):
        items.append({
            "feature": feature, "label": label, "action": action,
            "priority": priority, "topic": topic, "consumed": False,
            "value": num(features.get(feature)),
        })

    past_falls = num(features.get("past_falls", 0) or 0)
    tug = num(features.get("tug_seconds"))
    mobility = num(features.get("mobility_score"))
    cognitive = features.get("cognitive_impairment")
    night_act = num(features.get("night_activity_duration_min"))
    bed_exits = num(features.get("night_bed_exits"))
    dslf = num(features.get("days_since_last_fall"))
    polypharmacy = num(features.get("polypharmacy_count"))
    sex = features.get("sex")
    is_male = isinstance(sex, str) and sex.upper().startswith("M")

    # ---------- Layer 1: _describe（含数值的动态判讀 + 直白可落地的英文护理动作） ----------
    if features.get("syncopal_fall") == 1:
        add("syncopal_fall", "Syncopal fall",
            "Syncopal fall: report to the nursing station immediately; assist every sit-to-stand and transfer; ask the doctor to arrange an ECG and cardiac evaluation - the fall may be due to heart or blood pressure problems", 1, "heart")
    if features.get("fall_cluster_30d") == 1:
        add("fall_cluster_30d", "2+ falls in 30 days",
            "2+ falls in 30 days: record the time and circumstances of each fall and report to the nursing station; look for new triggers (new medication, fever, dehydration); if the cause is unclear within 2 days, arrange clinic/ER referral", 1, "fall_workup")
    if past_falls >= 2:
        add("past_falls", f"{past_falls} falls this year",
            "Complete a full check this week: standing blood pressure, vision, feet, and all medications; because of 2 falls this year, also ask the doctor to refer to a falls clinic/geriatric assessment", 1, "fall_workup")
    if isinstance(tug, (int, float)):
        if 8 <= tug < 12:
            add("tug_seconds", f"TUG {tug}s (sensitive band)",
                "Start strength and balance training now: 3x/week, 30 min - sit-to-stand x10, heel-to-toe walking, single-leg stand holding a chair", 1, "training")
        elif 12 <= tug < 13.5:
            add("tug_seconds", f"TUG {tug}s (approaching high risk)",
                "Continue strength and balance training 3x/week; also ask a physiotherapist to assess the gait", 1, "training")
        elif 13.5 <= tug < 20:
            add("tug_seconds", f"TUG {tug}s (high risk)",
                "Stop routine training; ask a physiotherapist to start a fall-prevention plan and assess whether a walking aid is needed", 1, "training")
        elif tug >= 20:
            add("tug_seconds", f"TUG {tug}s (severe)",
                "Use 2-person assist or a transfer aid for all transfers; ask an occupational therapist to assess bed-side and bathroom transfer routes", 1, "training")
    if isinstance(mobility, (int, float)) and 1 <= mobility <= 4:
        add("mobility_score", f"Mobility {mobility}/10 (impaired)",
            "Assist with leg/balance training 3x/week; if no improvement in 6 weeks, refer to a physiotherapist; consider an assistive device if walking is unsteady", 1, "training")
    if features.get("high_risk_medication") == 1:
        add("high_risk_medication", "Fall-risk medication",
            "List all medications and give to the nurse: ask the doctor to review fall-risk drugs (sleeping pills, antidepressants, painkillers); never stop them on your own", 1, "med")
    if features.get("orthostatic_hypotension") == 1:
        add("orthostatic_hypotension", "Dizzy on standing (orthostatic hypotension)",
            "Teach the 3-step rise: lie 30s - sit 30s - stand 30s before walking; 6-8 glasses of water a day; ask the doctor to review blood pressure medication", 1, "bp")
    if past_falls >= 1 and isinstance(dslf, (int, float)) and 0 <= dslf < 30:
        add("days_since_last_fall", f"Fell {dslf} days ago (recent)",
            "Fell within the last 30 days: place a 'recent fall' sign at the bedside; accompany all bed exits; observe every hour for 24 hours after the fall to prevent a second fall", 1, "fall_workup")

    if past_falls == 1:
        add("past_falls", "1 fall this year",
            "1 fall this year: complete a full check within 2 weeks - standing blood pressure, vision, feet, and all medications", 2, "fall_workup")
    if cognitive == 1:
        add("cognitive_impairment", "Mild cognitive impairment",
            "Mild cognitive impairment: at a fixed time daily, do simple activities with the resident (10-min walk, 5 sit-to-stands); if more confused or drowsy than usual, report to the nursing station immediately", 2, "supervision")
    elif cognitive == 2:
        add("cognitive_impairment", "Moderate-to-severe cognitive impairment",
            "Moderate-to-severe cognitive impairment: one-to-one supervision for daily activities; keep corridors well-lit and clutter-free; record and report nighttime wandering or agitation", 2, "supervision")
    if isinstance(polypharmacy, (int, float)) and polypharmacy >= 5:
        add("polypharmacy_count", f"{polypharmacy} medications",
            "5+ medications: list all medications and give to the nurse; arrange a full medication review by a pharmacist", 2, "med")
    if isinstance(night_act, (int, float)) and night_act > 30:
        add("night_activity_duration_min", f"Awake {night_act} min at night",
            "Awake more than 30 min at night: record wake times and duration; check for pain, temperature discomfort, or loud snoring (possible sleep problem); no fluids 2h before bed; if frequently awake, report to the nursing station", 2, "sleep")

    if isinstance(bed_exits, (int, float)) and bed_exits >= 2:
        if bed_exits == 2:
            add("night_bed_exits", "2 bed exits per night",
                "2 bed exits per night: no fluids 2h before bed; keep a night light on; place non-slip slippers by the bed and put them on before getting up", 3, "night_safety")
        else:
            organ = "prostate (male)" if is_male else "bladder (female)"
            add("night_bed_exits", f"{bed_exits} bed exits per night (frequent)",
                f"3+ bed exits per night: place a commode by the bed to shorten nighttime walking; keep a night light on; clear clutter from the bedside; record the nightly exit count; ask the doctor to check {organ} issues", 3, "night_safety")

    # ---------- Layer 2: _apply_combos（组合吸收，被吸收条打 consumed） ----------
    def find(cond):
        for it in items:
            if not it["consumed"] and cond(it):
                return it
        return None

    def combo(feature, label, action, priority, topic):
        items.append({
            "feature": feature, "label": label, "action": action,
            "priority": priority, "topic": topic, "consumed": False, "value": None,
        })

    # 1) 起身暈 + 步態/行動力問題 → 一条 P1：先處理血壓再訓練
    bp = find(lambda it: it["feature"] == "orthostatic_hypotension")
    train_signals = [it for it in items if not it["consumed"] and it["feature"] in ("tug_seconds", "mobility_score")]
    if bp and train_signals:
        bp["consumed"] = True
        main = train_signals[0]
        main["consumed"] = True
        for t in train_signals[1:]:
            t["consumed"] = True
        combo("orthostatic+tug", "Dizziness + gait/mobility problems",
              "Dizzy on standing with slow gait: teach the 3-step rise (lie 30s - sit 30s - stand 30s) and check standing blood pressure daily; start balance training 3x/week only after blood pressure is stable", 1, "training")

    # 2) 跌倒≥2 + 暈厥 → 一条 P1：心臟科
    pf = find(lambda it: it["feature"] == "past_falls" and it["priority"] == 1)
    sync = find(lambda it: it["feature"] == "syncopal_fall")
    if pf and sync:
        pf["consumed"] = True
        sync["consumed"] = True
        combo("past_falls+syncopal", "Fall history + syncope",
              "2 falls this year plus a syncopal fall: report to the nursing station immediately; ask the doctor to prioritize cardiac evaluation (ECG, blood pressure)", 1, "heart")

    # 3) 認知障礙 + 夜間清醒>30 → 一条 P2：夜間專人巡視
    cog = find(lambda it: it["feature"] == "cognitive_impairment")
    na = find(lambda it: it["feature"] == "night_activity_duration_min")
    if cog and na:
        cog["consumed"] = True
        na["consumed"] = True
        combo("cognitive+night_activity", "Cognitive impairment + nighttime wakefulness",
              "Cognitive impairment with nighttime wakefulness: check on the resident every 2 hours at night, record wakefulness and activity; report unusual behavior immediately", 2, "supervision")

    # 4) TUG + 行動力低 → 合并一条 P1 training（无起身暈时的训练合并）
    tug_sig = find(lambda it: it["feature"] == "tug_seconds")
    mob_sig = find(lambda it: it["feature"] == "mobility_score")
    if tug_sig and mob_sig:
        tug_sig["consumed"] = True
        mob_sig["consumed"] = True
        tv, mv = tug_sig["value"], mob_sig["value"]
        if tv >= 13.5:
            action = "High-risk gait with low mobility: stop routine training; ask a physiotherapist to assess and design a training plan; consider an assistive device"
        else:
            action = "Slow gait with low mobility: strength and balance training 3x/week, 30 min - sit-to-stand x10, heel-to-toe walking, single-leg stand holding a chair; accompany walking"
        combo("tug+mobility", f"TUG {tv}s + Mobility {mv}/10", action, 1, "training")

    # 5) 風險藥物 + 起身暈 → 一条 P1：合併審查血壓藥
    med = find(lambda it: it["feature"] == "high_risk_medication")
    bp2 = find(lambda it: it["feature"] == "orthostatic_hypotension")
    if med and bp2:
        med["consumed"] = True
        bp2["consumed"] = True
        combo("medication+orthostatic", "Fall-risk drugs + dizziness",
              "Taking fall-risk drugs and dizzy on standing: list all medications and give to the nurse; ask the doctor to review blood pressure drugs and fall-risk drugs together", 1, "med")

    # 6) 多重用藥≥5 + 風險藥物 → 一条 P2：藥師審全部藥
    poly = find(lambda it: it["feature"] == "polypharmacy_count")
    med2 = find(lambda it: it["feature"] == "high_risk_medication")
    if poly and med2:
        poly["consumed"] = True
        med2["consumed"] = True
        combo("polypharmacy+medication", "Polypharmacy + fall-risk drugs",
              "5+ medications including fall-risk drugs: list all medications and give to the nurse; arrange a full pharmacist review", 2, "med")

    # 7) 跌倒史 + 近期(<30 天) → 一条 P1：防二次跌倒
    pf2 = find(lambda it: it["feature"] == "past_falls")
    d2 = find(lambda it: it["feature"] == "days_since_last_fall")
    if pf2 and d2:
        pf2["consumed"] = True
        d2["consumed"] = True
        combo("past_falls+recent", f"{pf2['value']} falls this year, recent fall",
              f"{pf2['value']} falls this year with a fall within 30 days: place a sign at the bedside, accompany all bed exits, 24-48h special supervision; arrange a full check as soon as possible", 1, "fall_workup")

    # ---------- Layer 3: _finalize（过滤 consumed + 排序 + 限 6 条） ----------
    out = [{"feature": it["feature"], "label": it["label"],
            "action": it["action"], "priority": it["priority"]}
           for it in items if not it["consumed"]]
    out.sort(key=lambda it: it["priority"])
    return out[:6]


def build_suggestions(features: dict) -> dict:
    """建议入口：仅 60-74 岁带输出建议；其他年龄输出 Not suggestion。"""
    age = features.get("age")
    if not isinstance(age, (int, float)) or not (60 <= age <= 74):
        return {"band": None, "not_suggestion": True, "items": []}
    return {"band": "60-74", "not_suggestion": False, "items": _suggest_60_74(features)}


app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.post("/token", response_model=Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends()):
    user = MOCK_USERS_DB.get(form_data.username)
    if not user or not verify_password(form_data.password, user["hashed_password"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["username"]}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}


@app.post("/predict")
async def get_prediction(data: PatientData, db: AsyncSession = Depends(get_db), current_user: dict = Depends(get_current_user)):
    try:
        # Convert incoming JSON data into a clean Python dictionary
        features_dict = data.model_dump()
        
        # Run model function
        result = predict_fall_risk(features_dict)
        lime_explanations = explain_patient(features_dict, max_features=3)
        suggestion = build_suggestions(features_dict)

        # Save feature profile & prediction output to database
        db_record = PatientRecord(
            sex=features_dict.get("sex"),
            age=features_dict["age"],
            night_bed_exits=features_dict["night_bed_exits"],
            night_activity_duration_min=features_dict["night_activity_duration_min"],
            past_falls=features_dict["past_falls"],
            mobility_score=features_dict["mobility_score"],
            high_risk_medication=features_dict["high_risk_medication"],
            cognitive_impairment=features_dict["cognitive_impairment"],
            polypharmacy_count=features_dict["polypharmacy_count"],
            orthostatic_hypotension=features_dict["orthostatic_hypotension"],
            tug_seconds=features_dict["tug_seconds"],
            days_since_last_fall=features_dict.get("days_since_last_fall"),
            syncopal_fall=features_dict.get("syncopal_fall", 0),
            fall_cluster_30d=features_dict.get("fall_cluster_30d", 0),
            fall_risk_level=result,
            resident_id=features_dict.get("resident_id"),
            lime_explanations=json.dumps(lime_explanations),
        )
        db.add(db_record)
        await db.commit()
        
        # Return the exact JSON structure your friend asked for
        return {
            "id": db_record.id,
            "fall_risk_level": result,
            "lime_explanations": lime_explanations,
            "suggestion": suggestion,
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Prediction failed: {str(e)}")
    
# ---------- Assessment query endpoints (for the front-end dashboard) ----------

def _record_to_dict(record: PatientRecord) -> dict:
    """Convert a PatientRecord row into a JSON-friendly dict."""
    return {
        "id": record.id,
        "sex": record.sex,
        "age": record.age,
        "night_bed_exits": record.night_bed_exits,
        "night_activity_duration_min": record.night_activity_duration_min,
        "past_falls": record.past_falls,
        "mobility_score": record.mobility_score,
        "high_risk_medication": record.high_risk_medication,
        "cognitive_impairment": record.cognitive_impairment,
        "polypharmacy_count": record.polypharmacy_count,
        "orthostatic_hypotension": record.orthostatic_hypotension,
        "tug_seconds": record.tug_seconds,
        "days_since_last_fall": record.days_since_last_fall,
        "syncopal_fall": record.syncopal_fall,
        "fall_cluster_30d": record.fall_cluster_30d,
        "fall_risk_level": record.fall_risk_level,
        "resident_id": record.resident_id,
        "lime_explanations": json.loads(record.lime_explanations) if record.lime_explanations else [],
        "suggestion": build_suggestions({
            "sex": record.sex,
            "age": record.age,
            "night_bed_exits": record.night_bed_exits,
            "night_activity_duration_min": record.night_activity_duration_min,
            "past_falls": record.past_falls,
            "mobility_score": record.mobility_score,
            "high_risk_medication": record.high_risk_medication,
            "cognitive_impairment": record.cognitive_impairment,
            "polypharmacy_count": record.polypharmacy_count,
            "orthostatic_hypotension": record.orthostatic_hypotension,
            "tug_seconds": record.tug_seconds,
            "days_since_last_fall": record.days_since_last_fall,
            "syncopal_fall": record.syncopal_fall,
            "fall_cluster_30d": record.fall_cluster_30d,
        }),
        "created_at": record.created_at.isoformat() if record.created_at else None,
    }


def _format_dt(iso_str):
    if not iso_str:
        return "—"
    try:
        return datetime.fromisoformat(iso_str).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return iso_str


def generate_assessment_pdf(record: dict) -> bytes:
    """Generate a printable PDF for a single assessment (for family / supervisor)."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        title=f"Fall Risk Assessment - {record.get('resident_id') or record['id']}",
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    body = []

    # Title
    body.append(Paragraph("Fall Risk Assessment Report", styles["Title"]))
    body.append(Spacer(1, 0.3 * cm))

    # Resident info
    risk_level = record.get("fall_risk_level", "UNKNOWN")
    risk_color = {
        "HIGH": colors.HexColor("#EA5455"),
        "MEDIUM": colors.HexColor("#FF9F43"),
        "LOW": colors.HexColor("#28C76F"),
    }.get(risk_level, colors.grey)

    info_data = [
        ["Resident ID:", record.get("resident_id") or f"#{record['id']}"],
        ["Sex / Age:", f"{record.get('sex', '—') or '—'} / {record.get('age', '—')}"],
        ["Assessed at:", _format_dt(record.get("created_at"))],
    ]
    info_tbl = Table(info_data, colWidths=[4 * cm, 12 * cm])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#555")),
    ]))
    body.append(info_tbl)
    body.append(Spacer(1, 0.5 * cm))

    # Big risk banner
    banner = Table([[f"{risk_level} RISK"]], colWidths=[16 * cm])
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), risk_color),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 28),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))
    body.append(banner)
    body.append(Spacer(1, 0.6 * cm))

    # Features
    body.append(Paragraph("Resident Profile", styles["Heading2"]))
    feat_data = [
        ["Feature", "Value"],
        ["Age", str(record.get("age", "—"))],
        ["Sex", record.get("sex", "—") or "—"],
        ["Night bed exits (per night)", str(record.get("night_bed_exits", "—"))],
        ["Night activity duration (min)", str(record.get("night_activity_duration_min", "—"))],
        ["Past falls (last year)", str(record.get("past_falls", "—"))],
        ["Mobility score (1-10, higher = better)", str(record.get("mobility_score", "—"))],
        ["High-risk medication", "Yes" if record.get("high_risk_medication") else "No"],
        ["Cognitive impairment (0/1/2)", str(record.get("cognitive_impairment", "—"))],
        ["Polypharmacy count", str(record.get("polypharmacy_count", "—"))],
        ["Orthostatic hypotension", "Yes" if record.get("orthostatic_hypotension") else "No"],
        ["TUG test (seconds)", str(record.get("tug_seconds", "—"))],
        ["Days since last fall", str(record.get("days_since_last_fall") if record.get("days_since_last_fall") is not None else "—")],
        ["Syncopal fall (loss of consciousness)", "Yes" if record.get("syncopal_fall") else "No"],
        ["Acute fall cluster (2+ in 30 days)", "Yes" if record.get("fall_cluster_30d") else "No"],
    ]
    feat_tbl = Table(feat_data, colWidths=[10 * cm, 6 * cm])
    feat_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    body.append(feat_tbl)
    body.append(Spacer(1, 0.5 * cm))

    # LIME explanations
    lime = record.get("lime_explanations") or []
    if lime:
        body.append(Paragraph("Why this risk level", styles["Heading2"]))
        for i, e in enumerate(lime, 1):
            cond = e.get("condition", "")
            weight = e.get("weight", 0)
            direction = e.get("direction", "")
            body.append(Paragraph(
                f"{i}. {cond} <font color='#888'>(weight {weight:.3f}, {direction})</font>",
                styles["Normal"],
            ))
        body.append(Spacer(1, 0.3 * cm))

    body.append(Spacer(1, 0.6 * cm))
    body.append(Paragraph(
        f"<i>Generated at {datetime.now().isoformat(timespec='seconds')} — "
        f"This report is produced by the AI Fall Risk Assessment System for informational purposes. "
        f"Care decisions should always be made by qualified medical professionals.</i>",
        styles["Italic"],
    ))

    doc.build(body)
    pdf = buf.getvalue()
    buf.close()
    return pdf


@app.get("/assessments/summary")
async def get_assessment_summary(db: AsyncSession = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Summary stats for the dashboard: total + count by risk level."""
    from sqlalchemy import func, select
    result = await db.execute(
        select(PatientRecord.fall_risk_level, func.count()).group_by(PatientRecord.fall_risk_level)
    )
    counts = {level: n for level, n in result.all()}
    total = sum(counts.values())
    return {
        "total": total,
        "high": counts.get("HIGH", 0),
        "medium": counts.get("MEDIUM", 0),
        "low": counts.get("LOW", 0),
    }


@app.get("/assessments")
async def list_assessments(
    page: int = 1,
    itemsPerPage: int = 10,
    risk_level: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Paginated list of assessment records, optionally filtered by risk level."""
    from sqlalchemy import func, select
    stmt = select(PatientRecord)
    count_stmt = select(func.count()).select_from(PatientRecord)
    if risk_level:
        stmt = stmt.where(PatientRecord.fall_risk_level == risk_level)
        count_stmt = count_stmt.where(PatientRecord.fall_risk_level == risk_level)
    stmt = stmt.order_by(PatientRecord.created_at.desc()).offset((page - 1) * itemsPerPage).limit(itemsPerPage)
    total = (await db.execute(count_stmt)).scalar() or 0
    rows = (await db.execute(stmt)).scalars().all()
    items = [_record_to_dict(r) for r in rows]
    return {"items": items, "total": total}


@app.get("/assessments/{record_id}")
async def get_assessment(record_id: int, db: AsyncSession = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Single assessment detail."""
    record = await db.get(PatientRecord, record_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Assessment not found")
    return _record_to_dict(record)


@app.get("/assessments/{record_id}/pdf")
async def download_assessment_pdf(record_id: int, db: AsyncSession = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Generate and download a printable PDF for this assessment."""
    record = await db.get(PatientRecord, record_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Assessment not found")
    data = _record_to_dict(record)
    pdf_bytes = generate_assessment_pdf(data)
    filename = f"assessment-{data.get('resident_id') or record_id}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.delete("/assessments/all")
async def delete_all_assessments(db: AsyncSession = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Delete all assessment records (one-click clear)."""
    from sqlalchemy import delete
    await db.execute(delete(PatientRecord))
    await db.commit()
    return {"deleted": True}


class BatchDeleteRequest(BaseModel):
    ids: list[int]


@app.post("/assessments/batch-delete")
async def batch_delete_assessments(req: BatchDeleteRequest, db: AsyncSession = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Delete multiple assessment records by id."""
    from sqlalchemy import delete
    if not req.ids:
        return {"deleted": 0}
    await db.execute(delete(PatientRecord).where(PatientRecord.id.in_(req.ids)))
    await db.commit()
    return {"deleted": len(req.ids)}


@app.delete("/assessments/{record_id}")
async def delete_assessment(record_id: int, db: AsyncSession = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Delete a single assessment record."""
    record = await db.get(PatientRecord, record_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Assessment not found")
    await db.delete(record)
    await db.commit()
    return {"deleted": record_id}


# Gradio Prediction Logic
async def predict_gradio(
        inputSex,
        inputAge, 
        inputPastFalls, 
        inputMobilityScore, 
        inputNightBedExits, 
        inputNightActivityDurationMin, 
        inputHighRiskMed, 
        inputCognitiveImpairment,
        inputPolypharmacyCount,
        inputOrthostaticHypotension,
        inputTugSeconds,
        inputDaysSinceLastFall,
        inputSyncopalFall,
        inputFallCluster30d
    ):
    profile = PatientData(
        sex=_to_sex(inputSex),
        age=int(inputAge),
        night_bed_exits=int(inputNightBedExits),
        night_activity_duration_min=int(inputNightActivityDurationMin),
        past_falls=int(inputPastFalls),
        mobility_score=int(inputMobilityScore),
        high_risk_medication=1 if inputHighRiskMed else 0,
        cognitive_impairment=int(inputCognitiveImpairment), 
        polypharmacy_count=int(inputPolypharmacyCount), 
        orthostatic_hypotension=1 if inputOrthostaticHypotension else 0, 
        tug_seconds=float(inputTugSeconds),
        days_since_last_fall=int(inputDaysSinceLastFall) if inputDaysSinceLastFall else None,
        syncopal_fall=1 if inputSyncopalFall else 0,
        fall_cluster_30d=1 if inputFallCluster30d else 0,
    )

    async with AsyncSessionLocal() as session:
        response = await get_prediction(data=profile, db=session)
        
    risk_level = response["fall_risk_level"]
    explanations = response["lime_explanations"]

    # Format output for the user interface textboxes
    output_text = f"Risk Level: {risk_level}\n"
    output_text += "primary factors:\n"
    for i, exp in enumerate(explanations, start=1):
        output_text += f" {i}. {exp['condition']} | Weight: {exp['weight']} ({exp['direction']})\n"
    return output_text

def random_gradio():
    df = pd.read_csv(os.path.join(DATA_DIR, "fall_risk_patients_2000_v2.csv"))
    X = df.drop(columns=["fall_risk_score", "fall_risk_level", "patient_id", "sex"])    # feature
    first_row = X.iloc[random.randint(0, len(df) - 1)].to_dict()
    return (
        first_row.get("age"),
        first_row.get("night_bed_exits"),
        first_row.get("night_activity_duration_min"),
        first_row.get("past_falls"),
        first_row.get("mobility_score"),
        first_row.get("high_risk_medication"),
        first_row.get("cognitive_impairment"),
        first_row.get("polypharmacy_count"),
        first_row.get("orthostatic_hypotension"),
        first_row.get("tug_seconds"),
        first_row.get("days_since_last_fall") or 0,
        bool(first_row.get("syncopal_fall")) if first_row.get("syncopal_fall") is not None else False,
        bool(first_row.get("fall_cluster_30d")) if first_row.get("fall_cluster_30d") is not None else False,
    )


# 中文列名 -> 英文特征名（养老院 Excel 模板用中文列名，后端自动映射）
COLUMN_MAP = {
    "姓名": "name", "性别": "sex", "年龄": "age",
    "夜间离床次数": "night_bed_exits",
    "夜间活动时长(分钟)": "night_activity_duration_min",
    "夜间活动时长": "night_activity_duration_min",
    "过去跌倒次数": "past_falls", "活动能力评分": "mobility_score",
    "是否使用高风险药物": "high_risk_medication",
    "认知障碍程度": "cognitive_impairment", "多重用药数量": "polypharmacy_count",
    "是否有体位性低血压": "orthostatic_hypotension",
    "起立行走测试(秒)": "tug_seconds", "起立行走测试": "tug_seconds",
    "距上次跌倒天数": "days_since_last_fall",
    "是否晕厥跌倒": "syncopal_fall", "跌倒时是否失去意识": "syncopal_fall",
    "30天内是否连续跌倒": "fall_cluster_30d", "30天内连续跌倒": "fall_cluster_30d",
}

FEATURE_CN = {
    "sex": "性别", "age": "年龄", "night_bed_exits": "夜间离床次数",
    "night_activity_duration_min": "夜间活动时长", "past_falls": "过去跌倒次数",
    "mobility_score": "活动能力评分", "high_risk_medication": "高风险药物",
    "cognitive_impairment": "认知障碍", "polypharmacy_count": "多重用药",
    "orthostatic_hypotension": "体位性低血压", "tug_seconds": "起立行走测试(秒)",
}


def _cn(text):
    for en, cn in FEATURE_CN.items():
        text = text.replace(en, cn)
    return text


def _to_sex(v):
    """男/女/M/F/male/female/1/0 -> 'M'/'F'"""
    if isinstance(v, (int, float)) and not pd.isna(v):
        return "M" if int(v) == 1 else "F"
    s = str(v).strip().lower()
    if s in ("男", "m", "male"):
        return "M"
    if s in ("女", "f", "female"):
        return "F"
    raise ValueError(f"性别无法识别：{v}（请填 男 或 女）")


def _to_bool(v):
    """是/否/有/无/yes/no/true/false/1/0 -> 1/0"""
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)) and not pd.isna(v):
        return int(v)
    s = str(v).strip().lower()
    if s in ("是", "有", "yes", "y", "true", "1"):
        return 1
    if s in ("否", "不是", "没有", "无", "no", "n", "false", "0"):
        return 0
    raise ValueError(f"无法识别：{v}（请填 是 或 否）")


def _build_excel_template(path):
    """Generate the Excel template (English column names + dropdowns)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fall Risk Assessment"

    headers = ["sex", "age", "night_bed_exits", "night_activity_duration_min",
               "past_falls", "mobility_score", "high_risk_medication",
               "cognitive_impairment", "polypharmacy_count",
               "orthostatic_hypotension", "tug_seconds",
               "days_since_last_fall", "syncopal_fall", "fall_cluster_30d"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")

    dv_sex = DataValidation(type="list", formula1='"Male,Female"', allow_blank=True)
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_cog = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
    for dv in (dv_sex, dv_yesno, dv_cog):
        ws.add_data_validation(dv)
    dv_sex.add("A2:A2000")
    dv_yesno.add("G2:G2000")
    dv_cog.add("H2:H2000")
    dv_yesno.add("J2:J2000")
    dv_yesno.add("M2:M2000")
    dv_yesno.add("N2:N2000")

    widths = {"A": 10, "B": 8, "C": 16, "D": 24, "E": 12, "F": 16,
              "G": 20, "H": 18, "I": 18, "J": 22, "K": 14,
              "L": 20, "M": 16, "N": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    wb.save(path)


def batch_predict_file(file):
    """Batch prediction: read an uploaded Excel(.xlsx) or CSV, predict each
    resident, and return a results table + summary.
    """
    if file is None:
        raise gr.Error("Please upload an Excel / CSV file first")

    required = ["sex", "age", "night_bed_exits", "night_activity_duration_min",
                "past_falls", "mobility_score", "high_risk_medication",
                "cognitive_impairment", "polypharmacy_count",
                "orthostatic_hypotension", "tug_seconds"]

    RANGE = {
        "age": (60, 100), "night_bed_exits": (0, 8),
        "night_activity_duration_min": (0, 120), "past_falls": (0, 5),
        "mobility_score": (1, 10), "high_risk_medication": (0, 1),
        "cognitive_impairment": (0, 2), "polypharmacy_count": (0, 14),
        "orthostatic_hypotension": (0, 1), "tug_seconds": (8.0, 31.9),
    }

    fname = file.name.lower()
    if fname.endswith((".xlsx", ".xlsm")):
        df = pd.read_excel(file.name, engine="openpyxl")
    elif fname.endswith(".csv"):
        df = None
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                df = pd.read_csv(file.name, encoding=enc)
                break
            except Exception:
                continue
        if df is None:
            raise gr.Error("Could not read this CSV. Please check the format.")
    else:
        raise gr.Error("Please upload a .xlsx or .csv file (download the template below)")

    df = df.rename(columns={k: v for k, v in COLUMN_MAP.items() if k in df.columns})

    missing = [c for c in required if c not in df.columns]
    if missing:
        raise gr.Error("Missing columns: " + ", ".join(missing) +
                       ". Please use the template below.")

    if "sex" in df.columns:
        df["sex"] = df["sex"].apply(lambda v: _to_sex(v) if not pd.isna(v) else v)
    for c in ("high_risk_medication", "orthostatic_hypotension"):
        df[c] = df[c].apply(lambda v: _to_bool(v) if not pd.isna(v) else v)
    for c in ("syncopal_fall", "fall_cluster_30d"):
        if c in df.columns:
            df[c] = df[c].apply(lambda v: _to_bool(v) if not pd.isna(v) else 0)

    to_save = []
    n_ok = 0
    n_err = 0
    for idx, r in df.iterrows():
        pid = f"P{idx + 1:05d}"
        try:
            feats = {}
            for c in required:
                v = r[c]
                if pd.isna(v):
                    raise ValueError(f"{c} is empty")
                if c in RANGE:
                    try:
                        fv = float(v)
                    except (TypeError, ValueError):
                        raise ValueError(f"{c} is not a number: {v}")
                    lo, hi = RANGE[c]
                    if not (lo <= fv <= hi):
                        raise ValueError(f"{c}={v} out of range {lo}~{hi}")
                feats[c] = v

            # Optional extended fall-detail fields (default if column absent).
            feats["days_since_last_fall"] = None
            feats["syncopal_fall"] = 0
            feats["fall_cluster_30d"] = 0
            if "days_since_last_fall" in df.columns and not pd.isna(r["days_since_last_fall"]):
                feats["days_since_last_fall"] = int(r["days_since_last_fall"])
            if "syncopal_fall" in df.columns and not pd.isna(r["syncopal_fall"]):
                feats["syncopal_fall"] = int(r["syncopal_fall"])
            if "fall_cluster_30d" in df.columns and not pd.isna(r["fall_cluster_30d"]):
                feats["fall_cluster_30d"] = int(r["fall_cluster_30d"])

            level = predict_fall_risk(feats)
            expl = explain_patient(feats, max_features=3)
            to_save.append(PatientRecord(
                sex=feats.get("sex"),
                age=feats["age"],
                night_bed_exits=feats["night_bed_exits"],
                night_activity_duration_min=feats["night_activity_duration_min"],
                past_falls=feats["past_falls"],
                mobility_score=feats["mobility_score"],
                high_risk_medication=feats["high_risk_medication"],
                cognitive_impairment=feats["cognitive_impairment"],
                polypharmacy_count=feats["polypharmacy_count"],
                orthostatic_hypotension=feats["orthostatic_hypotension"],
                tug_seconds=feats["tug_seconds"],
                days_since_last_fall=feats.get("days_since_last_fall"),
                syncopal_fall=feats.get("syncopal_fall", 0),
                fall_cluster_30d=feats.get("fall_cluster_30d", 0),
                fall_risk_level=level,
                resident_id=pid,
                lime_explanations=json.dumps(expl),
            ))
            n_ok += 1
        except Exception:
            n_err += 1

    if to_save:
        async def _save():
            async with AsyncSessionLocal() as session:
                session.add_all(to_save)
                await session.commit()
        asyncio.run(_save())

    msg = (f"### ✅ Batch processed\n\n"
           f"- Successfully saved: **{n_ok}** residents\n"
           f"- Data errors skipped: **{n_err}** rows\n\n"
           f"Please go to the **Risk Dashboard** to view the results:\n\n"
           f"👉 [Open Risk Dashboard](http://localhost:5173/dashboards/fall-risk-dashboard)")
    if n_err:
        msg += ("\n\n> Some rows were skipped due to invalid data. "
                "Please check the values are within range and not empty.")
    return msg


# Build Gradio UI (single Blocks with two Tabs: single + batch)
with gr.Blocks() as interface:
    gr.Markdown("# 跌倒风险评估系统")
    TEMPLATE_PATH = os.path.join(DATA_DIR, "import_template.xlsx")
    if not os.path.exists(TEMPLATE_PATH):
        _build_excel_template(TEMPLATE_PATH)

    with gr.Tab("单条预测"):
        with gr.Row():
            with gr.Column():
                inputSex = gr.Radio(
                    choices=["男", "女"], 
                    label="性别",
                    value="男"
                )
                inputAge = gr.Number(
                    minimum=60, 
                    maximum=100, 
                    label="年龄", 
                    value=65
                )
                inputPastFalls = gr.Number(
                    minimum=0, 
                    label="过去跌倒次数", 
                    value=0
                )
                inputMobilityScore = gr.Slider(
                    minimum=1, 
                    maximum=10, 
                    step=1, 
                    label="活动能力评分", 
                    value=5
                )
                inputNightBedExits = gr.Number(
                    minimum=0, 
                    label="夜间离床次数",
                    value=0
                )
                inputNightActivityDurationMin = gr.Number(
                    minimum=0, 
                    label="夜间活动时长(分钟)",
                    value=0
                )
                inputHighRiskMed = gr.Checkbox(
                    label="高风险药物"
                )
                inputCognitiveImpairment = gr.Number(
                    minimum=0, 
                    maximum=2, 
                    label="认知障碍程度", 
                    value=0
                )
                inputPolypharmacyCount = gr.Number(
                    minimum=0, 
                    label="多重用药数量",
                    value=0
                )
                inputOrthostaticHypotension = gr.Checkbox(
                    label="体位性低血压"
                )
                inputTugSeconds = gr.Number(
                    minimum=8, 
                    label="起立行走测试(秒)", 
                    value=8.0
                )
                inputDaysSinceLastFall = gr.Number(
                    minimum=0,
                    label="距上次跌倒天数（0=没跌过）",
                    value=0
                )
                inputSyncopalFall = gr.Checkbox(
                    label="跌倒时是否失去意识（晕厥）"
                )
                inputFallCluster30d = gr.Checkbox(
                    label="30天内是否连续跌倒≥2次"
                )
                random_btn = gr.Button("随机生成示例老人资料")
                submit_btn = gr.Button("提交预测")

            with gr.Column():
                output = gr.Textbox(label="预测结果", lines=4)

        random_btn.click(
            fn=random_gradio, 
            inputs=[], 
            outputs=[
                inputAge, inputNightBedExits, inputNightActivityDurationMin, 
                inputPastFalls, inputMobilityScore, inputHighRiskMed, inputCognitiveImpairment,
                inputPolypharmacyCount, inputOrthostaticHypotension, inputTugSeconds,
                inputDaysSinceLastFall, inputSyncopalFall, inputFallCluster30d
            ]
        )

        submit_btn.click(
            fn=predict_gradio, 
            inputs=[
                inputSex, inputAge, inputPastFalls, inputMobilityScore, inputNightBedExits, 
                inputNightActivityDurationMin, inputHighRiskMed, inputCognitiveImpairment,
                inputPolypharmacyCount, inputOrthostaticHypotension, inputTugSeconds,
                inputDaysSinceLastFall, inputSyncopalFall, inputFallCluster30d
            ], 
            outputs=output
        )

    with gr.Tab("Batch Prediction"):
        gr.Markdown("""
# Batch Fall Risk Assessment

**Three steps:**

1. Click "Download Excel Template" below and open it in Excel
2. Fill in the data (one resident per row)
3. Save, upload above, and click "Run Batch Prediction"

**Column guide:**

| Column | How to fill |
|--------|-------------|
| sex | Select "Male" or "Female" |
| age | Number, 60–100 |
| night_bed_exits | Number, 0–8 |
| night_activity_duration_min | Number, 0–120 |
| past_falls | Number, 0–5 |
| mobility_score | Number, 1–10 (higher = better mobility) |
| high_risk_medication | Select "Yes" or "No" |
| cognitive_impairment | 0 / 1 / 2 (0=none, 1=mild, 2=moderate/severe) |
| polypharmacy_count | Number, 0–14 |
| orthostatic_hypotension | Select "Yes" or "No" |
| tug_seconds | Number, 8–31.9 (seconds) |
| days_since_last_fall | Number, 0–365 (empty if past_falls=0) |
| syncopal_fall | Select "Yes" or "No" (fall with loss of consciousness) |
| fall_cluster_30d | Select "Yes" or "No" (≥2 falls within 30 days) |

> Values out of range are marked as "Data Error". Please fill within range.
""")
        gr.DownloadButton(label="Download Excel Template", value=TEMPLATE_PATH)
        batch_file = gr.File(label="Upload filled Excel file (.xlsx)")
        batch_btn = gr.Button("Run Batch Prediction", variant="primary")
        batch_summary = gr.Markdown()
        batch_btn.click(fn=batch_predict_file, inputs=batch_file,
                        outputs=[batch_summary])

# Mount Gradio inside FastAPI application context properly
app = gr.mount_gradio_app(app, interface, path="/")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)